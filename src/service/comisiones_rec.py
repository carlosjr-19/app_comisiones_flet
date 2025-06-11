import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as Img
from openpyxl.styles import numbers
from service.env import MVNOS

def limpiar_archivo(csv, csv2):

    print(f"Total en CSV Marca: {len(csv)}")
    print(f"Total en Recargas General: {len(csv2)}")

    # Limpiar las columnas de precios: quitar $ y ,
    cols_a_limpiar = ['mvno_package_price', 'reference_price']

    for col in cols_a_limpiar:
        csv.loc[:, col] = csv[col].astype(str).str.replace(r'[\$,]', '', regex=True)
        csv[col] = pd.to_numeric(csv[col], errors='coerce')

    total_mpp = csv['mvno_package_price'].sum()
    total_rp = csv['reference_price'].sum()

    #validar si el mismo monto
    print(total_mpp == total_rp)

    print("mvno_package_price:", total_mpp)
    print("reference_price", total_rp)

    # Listar los números únicos en cada archivo
    msisdn_csv = set(csv['msisdn'])
    msisdn_csv2 = set(csv2['msisdn'])

    # Detectar líneas que están en uno y no en otro
    en_csv_no_en_csv2 = msisdn_csv - msisdn_csv2
    en_csv2_no_en_csv = msisdn_csv2 - msisdn_csv

    print(f"Líneas en csv Marca pero no en Recargas General: {en_csv_no_en_csv2}")
    print(f"Líneas en Recargas General pero no en csv Marca: {en_csv2_no_en_csv}")

    # Si quieres también ver diferencias en el conteo de recargas por línea
    conteo_csv = csv['msisdn'].value_counts()
    conteo_csv2 = csv2['msisdn'].value_counts()

    # Comparar recargas por línea
    comparacion = pd.DataFrame({
        'Reporte Marca': conteo_csv,
        'Reporte General': conteo_csv2
    }).fillna(0)

    # Filtrar solo donde difiere el número de recargas
    diferencias = comparacion[comparacion['Reporte Marca'] != comparacion['Reporte General']]

    print("Diferencias en número de recargas por línea:")
    print(diferencias)

    return csv, diferencias

def procesar_comisiones(df, comision_sales, comision, fecha):

    # Ordenar por fecha ascendente antes de agregar la fila TOTAL
    df = df.sort_values(by='date', ascending=True)

    # Crear la columna '20%' inicializada en 0
    df['comisión'] = 0.0

    # Crear la columna 'Bonificación fija $' y bono $ en el caso que la marca tenga precios distintos 
    df['bonificación fija $'] = 0.0
    df['bono $'] = 0.0

    porcentaje_comision = 0

    if comision == "20%":
        porcentaje_comision = 0.2
    elif comision == "15%":
        porcentaje_comision = 0.15

    precios_iguales = True

    if (df['mvno_package_price'] != df['reference_price']).all():
        precios_iguales = False
    
    # Calcular 20% solo donde los precios coincidan
    df.loc[df['mvno_package_price'] == df['reference_price'], 'comisión'] = (
        df['mvno_package_price'] * porcentaje_comision
    )

    #Calcular comisión cuando tienen diferentes precios
    df.loc[df['mvno_package_price'] != df['reference_price'], 'comisión'] = (
        (df['reference_price'] * porcentaje_comision) + (df['mvno_package_price'] - df['reference_price'])
    )

    df.loc[df['mvno_package_price'] != df['reference_price'], 'bono $'] = (
        (df['reference_price'] * porcentaje_comision)
    )

    df.loc[df['mvno_package_price'] != df['reference_price'], 'bonificación fija $'] = (
        (df['mvno_package_price'] - df['reference_price'])
    )

    # Aplicar condición para los que NO son 'Sales'
    condicion = df['channel'] != 'Sales'
    df.loc[condicion, 'transacción 4.14%'] = df.loc[condicion, 'mvno_package_price'] * 0.0414
    df.loc[condicion, 'tasa fija 3.65'] = 3.65

    # Convertir a numérico (por si acaso)
    df['comisión'] = pd.to_numeric(df['comisión'], errors='coerce')
    df['transacción 4.14%'] = pd.to_numeric(df['transacción 4.14%'], errors='coerce')
    df['tasa fija 3.65'] = pd.to_numeric(df['tasa fija 3.65'], errors='coerce')

    # Reemplazar NaN por 0 donde no aplica
    df[['transacción 4.14%', 'tasa fija 3.65']] = df[['transacción 4.14%', 'tasa fija 3.65']].fillna(0)

    # Redondear las columnas a 2 decimales
    df['transacción 4.14%'] = df['transacción 4.14%'].round(2)
    df['tasa fija 3.65'] = df['tasa fija 3.65'].round(2)
    df['comisión'] = df['comisión'].round(2)

    # Crear columna de comisión total (ya con los valores redondeados)
    df['comisión_total'] = (df['comisión'] - df['transacción 4.14%'] - df['tasa fija 3.65']).round(2)

    df['mes'] = fecha
    df['porcentaje'] = comision

    if comision_sales == "NO":
        #condicion cuando no se pagan comisiones por sales
        condicion = df['channel'] == 'Sales'
        df.loc[condicion, 'comisión_total'] = 0.0

        #condicion cuando no se pagan comisiones por sales
        condicion = df['channel'] == 'Sales'
        df.loc[condicion, 'comisión'] = 0.0

        #condicion cuando no se pagan comisiones por sales
        condicion = df['channel'] == 'Sales'
        df.loc[condicion, 'porcentaje'] = ""

        #condicion cuando no se pagan comisiones por sales bono $ es 0
        condicion = df['channel'] == 'Sales'
        df.loc[condicion, 'bono $'] = 0.0

        #condicion cuando no se pagan comisiones por sales bonificación fija $ es 0
        condicion = df['channel'] == 'Sales'
        df.loc[condicion, 'bonificación fija $'] = 0.0

    cols_a_limpiar = ['bono $', 'bonificación fija $']

    for col in cols_a_limpiar:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Calcular total
    total_comisiones = df['comisión_total'].sum()
    total_mpp = df['mvno_package_price'].sum()
    total_bono = df['comisión'].sum()
    total_transaccion = df['transacción 4.14%'].sum()
    total_desc_fijo = df['tasa fija 3.65'].sum()
    total_bono20 = df['bono $'].sum()
    total_bono_fijo = df['bonificación fija $'].sum()


    df['msisdn'] = df['msisdn'].astype(str)

    # Mostrar primeras filas para validar
    #print(df[['msisdn','channel' ,'mvno_name', 'reference_price', 'comisión', 'transacción 4.14%', 'tasa fija 3.65', 'comisión_total']].head(50))
    
    
    if precios_iguales:

        # Crear un DataFrame con la fila del total
        fila_total = pd.DataFrame({
            'mes': '',
            'porcentaje': '',
            'mvno_package_name': 'TOTAL',
            'mvno_package_price': [total_mpp],
            'comisión': [total_bono],
            'transacción 4.14%': [total_transaccion],
            'tasa fija 3.65': [total_desc_fijo],
            'comisión_total': [total_comisiones],
        })

        # Concatenar al DataFrame original
        df = pd.concat([df, fila_total], ignore_index=True)

        return df[['mvno_name','msisdn','channel','profile_sim','store_name','user_staff_name','transaction_id','date','mes','mvno_package_name','reference_price' 'mvno_package_price', 'porcentaje', 'comisión', 'transacción 4.14%', 'tasa fija 3.65', 'comisión_total']], precios_iguales
    
    else:
        # Crear un DataFrame con la fila del total
        fila_total = pd.DataFrame({
            'mes': '',
            'porcentaje': '',
            'reference_price': 'TOTAL',
            'mvno_package_price': [total_mpp],
            'bono $': [total_bono20],
            'bonificación fija $': [total_bono_fijo],
            'comisión': [total_bono],
            'transacción 4.14%': [total_transaccion],
            'tasa fija 3.65': [total_desc_fijo],
            'comisión_total': [total_comisiones],
        })

        # Concatenar al DataFrame original
        df = pd.concat([df, fila_total], ignore_index=True)

        return df[['mvno_name','msisdn','channel','profile_sim','store_name','user_staff_name','transaction_id','date','mes','mvno_package_name', 'reference_price','mvno_package_price', 'porcentaje', 'bono $', 'bonificación fija $', 'comisión', 'transacción 4.14%', 'tasa fija 3.65', 'comisión_total']], precios_iguales


def estilos_excel(df, marca, precios_iguales, fecha):
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "RECARGAS"

    # Insertar imagen 
    logo = Img('src/assets/images/logo.png')
    logo.anchor = 'A1'  # Esquina superior derecha
    ws.add_image(logo)

    # Encabezado "ACTIVACIONES" centrado
    ws.merge_cells('B5:L5')
    ws['B5'] = "RECARGAS"
    ws['B5'].font = Font(size=16, bold=True)
    ws['B5'].alignment = Alignment(horizontal='center')

    # Añadir columna de numeración
    df.insert(0, 'N#', range(1, len(df) + 1))

    # Sobreescribir la columna con el valor formateado
    df["mvno_package_price"] = df["mvno_package_price"].replace(r'[\$,]', '', regex=True).astype(float)
    df["comisión"] = df["comisión"].replace(r'[\$,]', '', regex=True).astype(float)
    df["transacción 4.14%"] = df["transacción 4.14%"].replace(r'[\$,]', '', regex=True).astype(float)
    df["tasa fija 3.65"] = df["tasa fija 3.65"].replace(r'[\$,]', '', regex=True).astype(float)
    df["comisión_total"] = df["comisión_total"].replace(r'[\$,]', '', regex=True).astype(float)


    # Convertir tu CSV a DataFrame para exportarlo
    df_final = df

    # Insertar datos a partir de la fila 7
    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), start=7):
        for c_idx, value in enumerate(row, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if precios_iguales == True:
                # Convertir a negativo si es una de las columnas que deben ser negativas
                # Las columnas P y Q son las 16 y 17 en Excel (ya que empiezas en columna 2)
                if c_idx in [16, 17] and r_idx > 7:  # r_idx > 7 para no afectar los encabezados
                    try:
                        if value is not None and str(value).strip() != '':  # Solo si hay un valor
                            value = -abs(float(value))
                    except (ValueError, TypeError):
                        pass  # Si no se puede convertir a número, dejamos el valor original
            else:
                # Convertir a negativo si es una de las columnas que deben ser negativas
                # Las columnas P y Q son las 16 y 17 en Excel (ya que empiezas en columna 2)
                if c_idx in [19, 20] and r_idx > 7:  # r_idx > 7 para no afectar los encabezados
                    try:
                        if value is not None and str(value).strip() != '':  # Solo si hay un valor
                            value = -abs(float(value))
                    except (ValueError, TypeError):
                        pass  # Si no se puede convertir a número, dejamos el valor original
            
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Formato de encabezados (fila 7)
            if r_idx == 7:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Borde negro
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Bordes para datos normales
            if r_idx > 7:
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # Centrar valores de columna N#  (columna 2 en Excel)
                if c_idx == 2:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Centrar valores de columna porcentaje (columna14 en Excel)
                if c_idx == 14:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # alinear a la derecha valores de columna  (columna 3, 15, 16, 17 y 18 en Excel)
                if c_idx == 13:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                if c_idx == 15:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                if c_idx == 16:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                if c_idx == 17:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                if c_idx == 18:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Ajustar altura de encabezado
        ws.row_dimensions[7].height = 30

        if precios_iguales == True:
            # leyenda "DESCUENTO" centrado en las celdas
            ws.merge_cells('P6:Q6')
            ws['P6'] = "DESCUENTO"
            ws['P6'].font = Font(size=12, bold=False)
            ws['P6'].alignment = Alignment(horizontal='center')
            ws['P6'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
            ws['P6'].border = thin_border
            thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
            ws['Q6'].border = thin_border


            #FORMATEAR COMO MONEDA después de haber insertado todo
            for col_letter in ['M', 'O', 'P', 'Q', 'R']:  # Ajusta letras según tus columnas reales
                for cell in ws[col_letter][7:]:  # desde fila 8 en adelante
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            #FORMATEAR CELDAS DE COLORES 
            for col_letter in ['P', 'Q']:  # Ajusta letras según tus columnas reales
                for cell in ws[col_letter][6:]:  # desde fila 7 en adelante
                    cell.fill = PatternFill(start_color="f9e79f", end_color="f9e79f", fill_type="solid")

            #FORMATEAR CELDAS DE COLORES 
            for col_letter in ['O', 'R']:  # Ajusta letras según tus columnas reales
                for cell in ws[col_letter][6:]:  # desde fila 7 en adelante
                    cell.fill = PatternFill(start_color="aed6f1", end_color="aed6f1", fill_type="solid")
        elif precios_iguales == False:
            # leyenda "DESCUENTO" centrado en las celdas
            ws.merge_cells('S6:T6')
            ws['S6'] = "DESCUENTO"
            ws['S6'].font = Font(size=12, bold=False)
            ws['S6'].alignment = Alignment(horizontal='center')
            ws['S6'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
            ws['S6'].border = thin_border
            thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
            ws['T6'].border = thin_border

            #FORMATEAR COMO MONEDA después de haber insertado todo
            for col_letter in ['M', 'N', 'P', 'Q', 'R', 'S', 'T', 'U']: 
                for cell in ws[col_letter][7:]:  # desde fila 8 en adelante
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            #FORMATEAR CELDAS DE COLORES AMARILLO
            for col_letter in ['S', 'T']:  # Ajusta letras según tus columnas reales
                for cell in ws[col_letter][6:]:  # desde fila 7 en adelante
                    cell.fill = PatternFill(start_color="f9e79f", end_color="f9e79f", fill_type="solid")

            #FORMATEAR CELDAS DE COLORES AZUL
            for col_letter in ['P', 'Q', 'R']:  # Ajusta letras según tus columnas reales
                for cell in ws[col_letter][6:]:  # desde fila 7 en adelante
                    cell.fill = PatternFill(start_color="aed6f1", end_color="aed6f1", fill_type="solid")


    # Ajustar ancho de columnas (le puedes personalizar los anchos)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Letra de columna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Antes de exportar, añade la columna de numeración
    df.insert(0, 'N°', range(1, len(df) + 1))

    # Cálculo total de comisiones
    total_comision = df['comisión_total'].sum()

    # Crear una fila de total con None o '' en campos no numéricos
    fila_total = [''] * len(df.columns)  # llena toda la fila con ''
    fila_total[df.columns.get_loc('mes')] = df['mes'].iloc[0]  # o lo que quieras mostrar en la columna 'mes'
    fila_total[df.columns.get_loc('comisión_total')] = total_comision  # coloca total en columna correspondiente

    # Añadir la fila al final del dataframe
    df.loc[len(df)] = fila_total

    #Colocar en blanco el header del excel
    ws.merge_cells('B1:R4')
    ws.merge_cells('M5:R5')

    # Guardar Excel
    wb.save(f"Comisiones_rec_{marca}_{fecha}.xlsx")


